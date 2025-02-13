from openpyxl import load_workbook
from lunardate import LunarDate
import datetime
import os

# 添加农历月份和日期的中文映射
LUNAR_MONTHS = {
    1: '正月', 2: '二月', 3: '三月', 4: '四月',
    5: '五月', 6: '六月', 7: '七月', 8: '八月',
    9: '九月', 10: '十月', 11: '冬月', 12: '腊月'
}

LUNAR_DAYS = {
    1: '初一', 2: '初二', 3: '初三', 4: '初四', 5: '初五',
    6: '初六', 7: '初七', 8: '初八', 9: '初九', 10: '初十',
    11: '十一', 12: '十二', 13: '十三', 14: '十四', 15: '十五',
    16: '十六', 17: '十七', 18: '十八', 19: '十九', 20: '二十',
    21: '廿一', 22: '廿二', 23: '廿三', 24: '廿四', 25: '廿五',
    26: '廿六', 27: '廿七', 28: '廿八', 29: '廿九', 30: '三十'
}

def convert_to_lunar(date):
    """将公历日期转换为农历日期"""
    lunar_date = LunarDate.fromSolarDate(date.year, date.month, date.day)
    return lunar_date

def process_excel(input_path, output_folder="output"):
    # 创建输出文件夹
    os.makedirs(output_folder, exist_ok=True)
    
    # 加载工作簿
    wb = load_workbook(input_path)
    ws = wb.active
    
    # 处理表头
    ws.cell(row=1, column=2, value="农历年份")
    ws.cell(row=1, column=3, value="农历月日")
    
    # 处理数据行
    for row in range(2, ws.max_row + 1):
        # 读取公历日期
        solar_date = ws.cell(row=row, column=1).value
        
        # 转换日期格式（新增对YYYYMMDD格式的支持）
        dt = None
        try:
            if isinstance(solar_date, datetime.datetime):
                dt = solar_date.date()
            else:
                # 处理数字格式或字符串格式的日期
                date_str = str(solar_date).strip()
                # 先尝试解析YYYYMMDD格式
                if len(date_str) == 8 and date_str.isdigit():
                    dt = datetime.datetime.strptime(date_str, "%Y%m%d").date()
                else:
                    # 尝试其他格式（如YYYY-MM-DD）
                    dt = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        except Exception as e:
            print(f"行 {row} 日期解析失败: {str(e)}")
            continue
        
        # 转换为农历日期
        try:
            lunar = convert_to_lunar(dt)
            # 写入农历年份
            ws.cell(row=row, column=2, value=f"{lunar.year}年")
            # 写入农历月日（中文格式）
            month_str = LUNAR_MONTHS.get(lunar.month, f"{lunar.month}月")
            day_str = LUNAR_DAYS.get(lunar.day, str(lunar.day))
            ws.cell(row=row, column=3, value=f"{month_str}{day_str}")
        except Exception as e:
            print(f"Error processing row {row}: {str(e)}")
            continue
    
    # 保存文件
    output_path = os.path.join(output_folder, f"processed_{os.path.basename(input_path)}")
    wb.save(output_path)
    print(f"文件已处理完成，保存至：{output_path}")

if __name__ == "__main__":
    # 使用示例（添加路径验证和错误处理）
    import os
    
    # 修改为实际文件路径（两种方案）
    input_file = os.path.join("file2", "日期处理.xlsx")  # 方案1：使用相对路径
    # input_file = os.path.abspath("file2/日期处理.xlsx")  # 方案2：使用绝对路径
    
    # 添加文件存在性检查
    if not os.path.exists(input_file):
        print(f"错误：文件 {input_file} 不存在！")
        print("当前工作目录：", os.getcwd())
        print("请确认：")
        print("1. 文件名是否完全一致（包括.xlsx扩展名）")
        print("2. 文件是否在以下路径：", os.path.abspath("file2"))
    else:
        process_excel(input_file) 