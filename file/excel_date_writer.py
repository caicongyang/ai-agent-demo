import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from pathlib import Path


class ExcelDateWriter:
    """
    A class for writing date data to Excel files.
    Creates Excel files with three columns:
    1. Daily dates from a start date to an end date
    2. Year and month of each date in YY.MM format (e.g., 24.01)
    3. Sequential number based on year-month (e.g., 24.01 = 1, 24.02 = 2, etc.)
    """

    def __init__(self, output_dir=None):
        """
        Initialize the ExcelDateWriter.
        
        Args:
            output_dir (str, optional): Directory to save the Excel file.
                                       If None, uses the current directory.
        """
        if output_dir is None:
            self.output_dir = os.getcwd()
        else:
            self.output_dir = output_dir
            # Create the directory if it doesn't exist
            Path(output_dir).mkdir(parents=True, exist_ok=True)
    
    def generate_dates(self, start_date, end_date):
        """
        Generate a list of dates from start_date to end_date (inclusive).
        
        Args:
            start_date (str): Start date in 'YYYY-MM-DD' format
            end_date (str): End date in 'YYYY-MM-DD' format
            
        Returns:
            list: List of datetime objects
        """
        # Convert string dates to datetime objects
        start = datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.strptime(end_date, '%Y-%m-%d')
        
        # Generate all dates in the range
        dates = []
        current_date = start
        while current_date <= end:
            dates.append(current_date)
            current_date += timedelta(days=1)
            
        return dates
    
    def format_year_month(self, date):
        """
        Format a date as 'YY.MM'.
        
        Args:
            date (datetime): Date to format
            
        Returns:
            str: Formatted date string
        """
        return date.strftime('%y.%m')  # Changed from '%Y.%m' to '%y.%m'
    
    def get_month_sequence_number(self, date, start_date):
        """
        Calculate the sequence number based on the difference in months from the start date.
        
        Args:
            date (datetime): Current date
            start_date (datetime): Start date
            
        Returns:
            int: Sequence number (starting from 1)
        """
        # Calculate the difference in months
        year_diff = date.year - start_date.year
        month_diff = date.month - start_date.month
        total_months = year_diff * 12 + month_diff + 1  # +1 to start from 1 instead of 0
        
        return total_months
    
    def write_to_excel(self, start_date='2024-01-01', end_date='2025-02-28', filename=None):
        """
        Write dates to an Excel file.
        
        Args:
            start_date (str): Start date in 'YYYY-MM-DD' format
            end_date (str): End date in 'YYYY-MM-DD' format
            filename (str, optional): Name of the Excel file.
                                     If None, generates a default name.
                                     
        Returns:
            str: Path to the created Excel file
        """
        # Generate dates
        dates = self.generate_dates(start_date, end_date)
        
        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Date Data"
        
        # Add headers
        ws['A1'] = "Date"
        ws['B1'] = "Year.Month"
        ws['C1'] = "Sequence Number"
        
        # Get start date as datetime for sequence calculation
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        
        # Track the current month to assign sequence numbers
        current_ym = None
        month_sequence = {}  # Dictionary to store year-month to sequence mapping
        
        # First pass: build the month sequence mapping
        seq_num = 1
        for date in dates:
            ym = self.format_year_month(date)
            if ym not in month_sequence:
                month_sequence[ym] = seq_num
                seq_num += 1
        
        # Add data
        for i, date in enumerate(dates, start=2):  # Start from row 2 (after header)
            ym = self.format_year_month(date)
            
            ws[f'A{i}'] = date.strftime('%Y-%m-%d')  # Date in YYYY-MM-DD format
            ws[f'B{i}'] = ym  # Year.Month in YY.MM format
            ws[f'C{i}'] = month_sequence[ym]  # Sequence number
        
        # Generate filename if not provided
        if filename is None:
            filename = f"dates_{start_date}_to_{end_date}.xlsx"
        
        # Ensure filename has .xlsx extension
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        # Save the workbook
        file_path = os.path.join(self.output_dir, filename)
        wb.save(file_path)
        
        print(f"Excel file created successfully at: {file_path}")
        return file_path


# Example usage
if __name__ == "__main__":
    writer = ExcelDateWriter()
    excel_file = writer.write_to_excel()
    print(f"Generated Excel file with dates from 2024-01-01 to 2025-02-28") 